from zipfile import BadZipFile
import openpyxl
from django.db import IntegrityError
from .models import Data


class Parser:
    def __init__(self, file, date_1, date_2):
        self.file = file
        self.date_1 = date_1
        self.date_2 = date_2

    @staticmethod
    def create_db_record(company_id, company_name, current_date, qliq, qoil, data_type):
        Data.objects.create(company_id=company_id, company_name=company_name, date=current_date,
                            qliq=qliq, qoil=qoil, data_type=data_type)

    def get_worksheet(self):
        workbook = openpyxl.load_workbook(self.file)
        return workbook.active

    def get_parsed_data(self):
        try:
            worksheet = self.get_worksheet()
        except BadZipFile:
            return {"error": "Ошибка при открытии файла. Проверьте его формат"}
        for row in range(4, worksheet.max_row + 1):
            company_id = worksheet.cell(row=row, column=1).value
            company_name = worksheet.cell(row=row, column=2).value
            for date_number in range(1, 3):
                for i in range(2, 7, 4):
                    qliq = worksheet.cell(row=row, column=date_number + i).value
                    qoil = worksheet.cell(row=row, column=date_number + i + 2).value
                    data_type = "0" if i == 2 else "1"
                    current_date = self.date_1 if date_number % 2 == 1 else self.date_2
                    try:
                        self.create_db_record(company_id, company_name, current_date, qliq, qoil, data_type)
                    except IntegrityError:
                        return {"error": "Ошибка при чтении файла. Аналогичные позиции уже есть в БД"}
                    except (ValueError, TypeError):
                        return {"error": "Ошибка при чтении файла. Некорректный формат данных"}
        return True
